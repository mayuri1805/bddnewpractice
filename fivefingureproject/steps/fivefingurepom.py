import openpyxl


username=''
password=''
classname=''
subjectname=''
bookseries=''
chaptername=''



path = "C:\\Users\Asus\\OneDrive\\Documents\\fivefingurepomfile.xlsx"


workbook = openpyxl.load_workbook(path)
sheet= workbook['Sheet1']

rows = sheet.max_row
cols = sheet.max_column
print(rows)
print(cols)

for i in range(1,rows):
    if sheet.cell(row=i+1,column=2).value == "fivefingurelogin":
        username = sheet.cell(row=i+1, column=3).value
        password = sheet.cell(row=i+1, column=4).value
        classname = sheet.cell(row=i+1, column=5).value
        subjectname = sheet.cell(row=i+1, column=6).value
        bookseries = sheet.cell(row=i + 1, column=7).value
        chaptername = sheet.cell(row=i + 1, column=8).value



        print(username,password,bookseries,chaptername)

